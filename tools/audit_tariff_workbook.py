from __future__ import annotations

import re
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"


def parse_excel_like_date(value):
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M %z", "%d/%m/%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:22], fmt).replace(tzinfo=None)
        except ValueError:
            pass
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2})", text)
    if m:
        day, month, year, hour, minute = map(int, m.groups())
        return datetime(year, month, day, hour, minute)
    return None


def code_for(dt: datetime, tarif: str) -> str:
    t = dt.time()
    wd = dt.weekday() + 1
    m = dt.month
    tarif = tarif or ""
    if tarif == "Vert TE":
        if m >= 10 or m <= 3:
            if wd >= 6:
                return "HCSH" if t < time(16, 0) else "HPSH"
            if t < time(8, 0):
                return "HCSH"
            if t < time(18, 0):
                return "HPSH"
            if t < time(22, 0):
                return "PO"
            return "HPSH"
        if wd >= 6:
            return "HCSB"
        return "HPSB" if time(18, 0) <= t < time(22, 0) else "HCSB"
    if tarif == "Vert":
        hiver = 5 <= m <= 9
        if t >= time(22, 30) or t < time(6, 30):
            return "HCH" if hiver else "HCE"
        if wd >= 6:
            return "HPH" if hiver else "HPE"
        if time(9, 0) <= t < time(12, 30) or time(19, 0) <= t < time(20, 30):
            return "PO"
        return "HPH" if hiver else "HPE"
    return "HC" if t >= time(22, 0) or t < time(6, 0) else "HP"


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
    print("sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"{name}: rows={ws.max_row} cols={ws.max_column}")

    ws = wb["Taitements"]
    print("selector:", ws["A4"].value, ws["B4"].value, ws["C4"].value)
    print("headers row 7:", [ws.cell(7, c).value for c in range(1, min(ws.max_column, 14) + 1)])
    print("first data row:", [ws.cell(8, c).value for c in range(1, 13)])

    sites = wb["Sites"]
    headers = [sites.cell(1, c).value for c in range(1, sites.max_column + 1)]
    print("site headers:", headers)
    unique_sites = [sites.cell(r, 20).value for r in range(2, sites.max_row + 1) if sites.cell(r, 20).value]
    print("unique site count:", len(unique_sites), "first:", unique_sites[:5])

    for needle in ["Aurar Omega", "Aurar St Benoit"]:
        for r in range(2, sites.max_row + 1):
            if sites.cell(r, 1).value == needle:
                print("site sample:", needle, r, [sites.cell(r, c).value for c in [1, 2, 3, 4, 5, 6, 10, 15, 16, 17, 18]])
                break

    tests = [
        (datetime(2025, 1, 1, 0, 0), "Vert TE"),
        (datetime(2025, 1, 1, 19, 0), "Vert TE"),
        (datetime(2025, 6, 2, 19, 0), "Vert TE"),
        (datetime(2025, 6, 2, 12, 0), "Vert TE"),
        (datetime(2025, 1, 1, 0, 0), "Bleu Plus"),
        (datetime(2025, 1, 1, 12, 0), "Bleu Plus"),
    ]
    for dt, tarif in tests:
        print("test_code:", dt, tarif, code_for(dt, tarif))


if __name__ == "__main__":
    main()
