from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
    ws = wb["Taitements"]
    sites = wb["Sites"]
    print("headers traitements", [ws.cell(7, col).value for col in range(1, 13)])
    print("A4/B4/C4", ws["A4"].value, ws["B4"].value, ws["C4"].value)
    print("E8", str(ws["E8"].value)[:160])
    print("I8", str(ws["I8"].value)[:220])
    print("liste sites", [sites.cell(row, 20).value for row in range(1, 8)])
    for row in range(2, sites.max_row + 1):
        if sites.cell(row, 1).value == "Aurar Omega":
            print("omega", row, [sites.cell(row, col).value for col in [1, 2, 3, 4, 6, 10, 15, 16, 17, 18]])
            break


if __name__ == "__main__":
    main()
