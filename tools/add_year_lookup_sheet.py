from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"

BLUE = "1F4E78"
GREEN = "70AD47"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
WHITE = "FFFFFF"


def find_header_row(ws, first_header: str) -> int:
    for row in range(1, min(ws.max_row, 20) + 1):
        if ws.cell(row, 1).value == first_header:
            return row
    raise ValueError(f"Header {first_header!r} not found in {ws.title}")


def unique_years(wb) -> list[int]:
    years = set()
    if "Prix" in wb.sheetnames:
        ws = wb["Prix"]
        header_row = find_header_row(ws, "date_application")
        headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
        if "annee" in headers:
            year_col = headers.index("annee") + 1
            for row in range(header_row + 1, ws.max_row + 1):
                value = ws.cell(row, year_col).value
                if isinstance(value, int):
                    years.add(value)
    if "Plages horaires" in wb.sheetnames:
        ws = wb["Plages horaires"]
        header_row = find_header_row(ws, "date_application")
        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row, 1).value
            if value:
                try:
                    years.add(int(str(value)[:4]))
                except ValueError:
                    pass
    return sorted(years)


def style_header_row(ws, row: int, start_col: int, end_col: int, fill: str) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build() -> None:
    wb = load_workbook(WORKBOOK)
    sheet_name = "Recherche année"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    years = unique_years(wb)
    selected_year = max(years) if years else 2026

    ws["A1"] = "Recherche par année"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Année à afficher"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = selected_year
    ws["B2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["B2"].alignment = Alignment(horizontal="center")
    if years:
        validation = DataValidation(type="list", formula1=f'"{",".join(str(y) for y in years)}"', allow_blank=False)
        validation.error = "Choisis une année présente dans la liste."
        validation.errorTitle = "Année non disponible"
        validation.prompt = "Choisis l'année à afficher."
        validation.promptTitle = "Filtre année"
        ws.add_data_validation(validation)
        validation.add(ws["B2"])

    ws["A4"] = "Prix liés à l'année choisie"
    ws["A4"].font = Font(bold=True, color=WHITE)
    ws["A4"].fill = PatternFill("solid", fgColor=GREEN)
    ws.merge_cells("A4:M4")

    prix = wb["Prix"]
    prix_header = find_header_row(prix, "date_application")
    prix_cols = prix.max_column
    prix_last = prix.max_row
    for col in range(1, prix_cols + 1):
        ws.cell(5, col, prix.cell(prix_header, col).value)
    style_header_row(ws, 5, 1, prix_cols, BLUE)
    prix_range = f"'Prix'!A{prix_header + 1}:{get_column_letter(prix_cols)}{prix_last}"
    year_range = f"'Prix'!B{prix_header + 1}:B{prix_last}"
    ws["A6"] = f'=FILTER({prix_range},{year_range}=$B$2,"Aucune donnée prix pour cette année")'

    plages_start_col = 15
    plages_title = ws.cell(4, plages_start_col, "Plages horaires liées à l'année choisie")
    plages_title.font = Font(bold=True, color=WHITE)
    plages_title.fill = PatternFill("solid", fgColor=GREEN)
    ws.merge_cells(start_row=4, start_column=plages_start_col, end_row=4, end_column=plages_start_col + 5)

    plages = wb["Plages horaires"]
    plages_header = find_header_row(plages, "date_application")
    plages_cols = plages.max_column
    plages_last = plages.max_row
    for col in range(1, plages_cols + 1):
        ws.cell(5, plages_start_col + col - 1, plages.cell(plages_header, col).value)
    style_header_row(ws, 5, plages_start_col, plages_start_col + plages_cols - 1, BLUE)
    plages_range = f"'Plages horaires'!A{plages_header + 1}:{get_column_letter(plages_cols)}{plages_last}"
    date_range = f"'Plages horaires'!A{plages_header + 1}:A{plages_last}"
    ws.cell(6, plages_start_col, f'=FILTER({plages_range},YEAR(DATEVALUE({date_range}))=$B$2,"Aucune plage horaire pour cette année")')

    widths = {
        "A": 16,
        "B": 12,
        "C": 18,
        "D": 30,
        "E": 24,
        "F": 20,
        "G": 16,
        "H": 18,
        "I": 24,
        "J": 20,
        "K": 20,
        "L": 18,
        "M": 20,
        "N": 4,
        "O": 16,
        "P": 16,
        "Q": 28,
        "R": 26,
        "S": 24,
        "T": 46,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False
    thin = Side(style="hair", color="DDDDDD")
    for row in ws.iter_rows(min_row=1, max_row=120, max_col=20):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    wb.save(WORKBOOK)

    check = load_workbook(WORKBOOK, data_only=False)
    assert sheet_name in check.sheetnames
    assert check[sheet_name]["B2"].value == selected_year
    assert str(check[sheet_name]["A6"].value).startswith("=FILTER(")
    print(WORKBOOK)
    print({"sheet": sheet_name, "annees": years, "annee_par_defaut": selected_year})


if __name__ == "__main__":
    build()
