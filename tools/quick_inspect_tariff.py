from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=False, data_only=False, keep_links=False)
    sites = wb["Sites"]
    print("Sites headers", [sites.cell(1, c).value for c in range(1, 21)])
    for row in [2, 99, 100, 452]:
        print("Sites row", row, [sites.cell(row, c).value for c in range(1, 19)])
    ws = wb["Taitements"]
    print("Taitements headers", [ws.cell(7, c).value for c in range(1, 16)])
    print("Taitements row8", [ws.cell(8, c).value for c in range(1, 16)])


if __name__ == "__main__":
    main()
