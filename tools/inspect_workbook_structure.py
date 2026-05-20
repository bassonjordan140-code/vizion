from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
    print(wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"--- {name} rows={ws.max_row} cols={ws.max_column}")
        for row in range(1, min(ws.max_row, 10) + 1):
            values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 24) + 1)]
            print(row, values)


if __name__ == "__main__":
    main()
