from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MAIN_FILE = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"
INVOICES_FILE = ROOT / "PythonProject" / "factures_edf.xlsx"

BLUE = "1F4E78"
GREEN = "70AD47"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
WHITE = "FFFFFF"
GREY = "F2F2F2"


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text[:10])
    except ValueError:
        return None


def parse_power_kva(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:[,.]\d+)?", str(value))
    return float(match.group(0).replace(",", ".")) if match else None


def normalize_tarif(value, power: float | None) -> str:
    text = (str(value or "")).strip().lower()
    if "vert" in text:
        return "Vert"
    if "bleu plus" in text or "bleu+" in text:
        return "Bleu Plus"
    if "bleu" in text:
        return "Bleu"
    if power is None:
        return ""
    if power >= 250:
        return "Vert"
    if power > 36:
        return "Bleu Plus"
    return "Bleu"


def load_sites() -> list[dict]:
    wb = load_workbook(INVOICES_FILE, data_only=True)
    ws = wb["Factures EDF"]
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    latest: dict[str, dict] = {}

    for row in range(3, ws.max_row + 1):
        site = ws.cell(row, idx["site"]).value
        if not site:
            continue
        period_end = parse_date(ws.cell(row, idx["periode_fin"]).value)
        power = parse_power_kva(ws.cell(row, idx["puissance_souscrite"]).value)
        tarif_raw = ws.cell(row, idx["tarif"]).value
        current = latest.get(site)
        if current is None or (period_end and current["periode_fin_ref"] and period_end > current["periode_fin_ref"]):
            latest[site] = {
                "site": str(site),
                "puissance_souscrite_kVA": power,
                "tarif_facture": tarif_raw,
                "tarif_retenu": normalize_tarif(tarif_raw, power),
                "periode_fin_ref": period_end,
            }

    return sorted(latest.values(), key=lambda item: item["site"].lower())


def recreate_sites_sheet(wb, sites: list[dict]) -> None:
    if "Sites" in wb.sheetnames:
        del wb["Sites"]
    ws = wb.create_sheet("Sites")
    headers = ["site", "puissance_souscrite_kVA", "tarif_facture", "tarif_retenu", "periode_fin_ref"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
    for row_idx, site in enumerate(sites, start=2):
        ws.cell(row_idx, 1, site["site"])
        ws.cell(row_idx, 2, site["puissance_souscrite_kVA"])
        ws.cell(row_idx, 3, site["tarif_facture"])
        ws.cell(row_idx, 4, site["tarif_retenu"])
        ws.cell(row_idx, 5, site["periode_fin_ref"])
        ws.cell(row_idx, 5).number_format = "yyyy-mm-dd"
    ref = f"A1:E{len(sites) + 1}"
    table = Table(displayName="Table_Sites", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    widths = [30, 22, 18, 16, 16]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def style_controls(ws) -> None:
    ws["D2"] = "Site sélectionné"
    ws["F2"] = "Puissance souscrite [kVA]"
    ws["H2"] = "Tarif retenu"
    for cell_ref in ("D2", "F2", "H2"):
        ws[cell_ref].font = Font(bold=True, color=WHITE)
        ws[cell_ref].fill = PatternFill("solid", fgColor=BLUE)
        ws[cell_ref].alignment = Alignment(horizontal="center")
    for cell_ref in ("E2", "G2", "I2"):
        ws[cell_ref].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws[cell_ref].alignment = Alignment(horizontal="center")
    ws["G2"] = '=IFERROR(VLOOKUP($E$2,Sites!$A:$E,2,FALSE),"")'
    ws["I2"] = '=IFERROR(VLOOKUP($E$2,Sites!$A:$E,4,FALSE),"")'


def add_treatments_logic(wb, sites: list[dict]) -> None:
    if "Taitements" not in wb.sheetnames:
        raise ValueError("La feuille 'Taitements' est introuvable.")
    ws = wb["Taitements"]
    max_row = ws.max_row

    style_controls(ws)
    if sites:
        ws["E2"] = sites[0]["site"]
        validation = DataValidation(type="list", formula1=f"=Sites!$A$2:$A${len(sites) + 1}", allow_blank=False)
        validation.error = "Choisis un site présent dans l'onglet Sites."
        validation.errorTitle = "Site non disponible"
        validation.prompt = "Choisis le site à analyser."
        validation.promptTitle = "Sélecteur de site"
        ws.add_data_validation(validation)
        validation.add(ws["E2"])

    headers = {
        "C3": "Date/heure calculée",
        "D3": "Site",
        "E3": "Puissance souscrite [kVA]",
        "F3": "Tarif retenu",
        "G3": "Saison",
        "H3": "Plage horaire",
        "I3": "Dépassement puissance",
        "J3": "Ecart vs puissance souscrite [kW]",
    }
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    plage_formula_template = (
        '=IF($F{r}="Vert",'
        'IF(OR(MOD($C{r},1)>=TIME(22,30,0),MOD($C{r},1)<TIME(6,30,0)),'
        'IF(AND(MONTH($C{r})>=5,MONTH($C{r})<=9),"Hiver - Heures Creuses","Eté - Heures Creuses"),'
        'IF(WEEKDAY($C{r},2)>=6,'
        'IF(AND(MONTH($C{r})>=5,MONTH($C{r})<=9),"Hiver - Heures Pleines","Eté - Heures Pleines"),'
        'IF(OR(AND(MOD($C{r},1)>=TIME(9,0,0),MOD($C{r},1)<TIME(12,30,0)),AND(MOD($C{r},1)>=TIME(19,0,0),MOD($C{r},1)<TIME(20,30,0))),'
        '"Pointe",'
        'IF(AND(MONTH($C{r})>=5,MONTH($C{r})<=9),"Hiver - Heures Pleines","Eté - Heures Pleines")))),'
        'IF(OR(MOD($C{r},1)>=TIME(22,0,0),MOD($C{r},1)<TIME(6,0,0)),"Heures Creuses","Heures Pleines"))'
    )

    for row in range(4, max_row + 1):
        ws.cell(row, 3).value = f'=IF($A{row}="","",DATE(VALUE(MID($A{row},7,4)),VALUE(MID($A{row},4,2)),VALUE(LEFT($A{row},2)))+TIME(VALUE(MID($A{row},12,2)),VALUE(MID($A{row},15,2)),0))'
        ws.cell(row, 4).value = '=$E$2'
        ws.cell(row, 5).value = '=$G$2'
        ws.cell(row, 6).value = '=$I$2'
        ws.cell(row, 7).value = f'=IF($C{row}="","",IF($F{row}="Vert",IF(AND(MONTH($C{row})>=5,MONTH($C{row})<=9),"Hiver austral","Eté austral"),"Toute l\'année"))'
        ws.cell(row, 8).value = plage_formula_template.format(r=row)
        ws.cell(row, 9).value = f'=IF(OR($B{row}="",$E{row}=""),"",IF($B{row}>$E{row},"Oui","Non"))'
        ws.cell(row, 10).value = f'=IF(OR($B{row}="",$E{row}=""),"",$B{row}-$E{row})'
        ws.cell(row, 3).number_format = "dd/mm/yyyy hh:mm"
        ws.cell(row, 5).number_format = "0.0"
        ws.cell(row, 10).number_format = "0.0"

    widths = {
        "A": 24, "B": 20, "C": 20, "D": 28, "E": 22,
        "F": 16, "G": 18, "H": 24, "I": 20, "J": 26,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    thin = Side(style="hair", color="DDDDDD")
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, 200), max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def main() -> None:
    sites = load_sites()
    wb = load_workbook(MAIN_FILE)
    recreate_sites_sheet(wb, sites)
    add_treatments_logic(wb, sites)
    wb.save(MAIN_FILE)

    check = load_workbook(MAIN_FILE, data_only=False, read_only=True)
    ws = check["Taitements"]
    assert "Sites" in check.sheetnames
    assert ws["E2"].value
    assert str(ws["H4"].value).startswith("=IF(")
    print(MAIN_FILE)
    print({"sites": len(sites), "lignes_pas_10min": ws.max_row - 3, "site_par_defaut": ws["E2"].value})


if __name__ == "__main__":
    main()
