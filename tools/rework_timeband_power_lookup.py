from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"

BLUE = "1F4E78"
GREEN = "70AD47"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"

SITE_COLS = {
    "date_ref_calc": 5,
    "tarif_retenu": 6,
    "P_BASE": 7,
    "P_HP": 8,
    "P_HC": 9,
    "P_PO": 10,
    "P_HPE": 11,
    "P_HCE": 12,
    "P_HPH": 13,
    "P_HCH": 14,
    "P_HPSH": 15,
    "P_HCSH": 16,
    "P_HPSB": 17,
    "P_HCSB": 18,
}


def parse_power_text(value) -> dict[str, float | None]:
    if value is None:
        return {}
    text = str(value).strip()
    pairs = re.findall(r"([A-Z]+)\s*:\s*(\d+(?:[,.]\d+)?)", text, flags=re.I)
    if pairs:
        return {code.upper(): float(num.replace(",", ".")) for code, num in pairs}
    match = re.search(r"\d+(?:[,.]\d+)?", text)
    if not match:
        return {}
    base = float(match.group(0).replace(",", "."))
    return {code: base for code in ["BASE", "HP", "HC", "PO", "HPE", "HCE", "HPH", "HCH", "HPSH", "HCSH", "HPSB", "HCSB"]}


def normalize_tarif(value) -> str:
    text = str(value or "").lower()
    if "vert" in text and ("te" in text or "transition" in text):
        return "Vert TE"
    if "vert" in text:
        return "Vert"
    if "bleu plus" in text or "bleu+" in text:
        return "Bleu Plus"
    if "bleu" in text:
        return "Bleu"
    return str(value or "")


def enhance_sites(ws) -> list[str]:
    headers = {
        1: "site",
        2: "puissance_souscrite_kVA",
        3: "tarif_facture",
        4: "periode_fin_ref",
        5: "date_ref_calc",
        6: "tarif_retenu",
        7: "P_BASE",
        8: "P_HP",
        9: "P_HC",
        10: "P_PO",
        11: "P_HPE",
        12: "P_HCE",
        13: "P_HPH",
        14: "P_HCH",
        15: "P_HPSH",
        16: "P_HCSH",
        17: "P_HPSB",
        18: "P_HCSB",
    }
    for col, header in headers.items():
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    sites: list[str] = []
    seen = set()
    for row in range(2, ws.max_row + 1):
        site = ws.cell(row, 1).value
        if site and site not in seen:
            sites.append(str(site))
            seen.add(site)
        powers = parse_power_text(ws.cell(row, 2).value)
        ws.cell(row, SITE_COLS["date_ref_calc"]).value = f'=IFERROR(DATEVALUE(D{row}),D{row})'
        ws.cell(row, SITE_COLS["tarif_retenu"]).value = normalize_tarif(ws.cell(row, 3).value)
        for code, col in [
            ("BASE", SITE_COLS["P_BASE"]),
            ("HP", SITE_COLS["P_HP"]),
            ("HC", SITE_COLS["P_HC"]),
            ("PO", SITE_COLS["P_PO"]),
            ("HPE", SITE_COLS["P_HPE"]),
            ("HCE", SITE_COLS["P_HCE"]),
            ("HPH", SITE_COLS["P_HPH"]),
            ("HCH", SITE_COLS["P_HCH"]),
            ("HPSH", SITE_COLS["P_HPSH"]),
            ("HCSH", SITE_COLS["P_HCSH"]),
            ("HPSB", SITE_COLS["P_HPSB"]),
            ("HCSB", SITE_COLS["P_HCSB"]),
        ]:
            ws.cell(row, col).value = powers.get(code) or powers.get("BASE")
    for col in range(1, 19):
        ws.column_dimensions[get_column_letter(col)].width = 20 if col != 1 else 30
    ws.cell(1, 20, "liste_sites")
    ws.cell(1, 20).font = Font(bold=True, color=WHITE)
    ws.cell(1, 20).fill = PatternFill("solid", fgColor=BLUE)
    for row, site in enumerate(sites, start=2):
        ws.cell(row, 20, site)
    ws.column_dimensions["T"].width = 30
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    return sites


def reset_site_validation(ws, sites_count: int) -> None:
    # Clear existing validations on the sheet to avoid stale list ranges.
    ws.data_validations.dataValidation = []
    validation = DataValidation(type="list", formula1=f"=Sites!$T$2:$T${sites_count + 1}", allow_blank=False)
    validation.error = "Choisis un site présent dans l'onglet Sites."
    validation.errorTitle = "Site non disponible"
    validation.prompt = "Choisis le site à analyser."
    validation.promptTitle = "Sélecteur de site"
    ws.add_data_validation(validation)
    validation.add(ws["A4"])


def add_treatments_formulas(ws, sites: list[str]) -> None:
    ws["A3"] = "Site sélectionné"
    ws["B3"] = "Puissance souscrite [kVA]"
    ws["C3"] = "Tarif retenu"
    for ref in ("A3", "B3", "C3"):
        ws[ref].font = Font(bold=True, color=WHITE)
        ws[ref].fill = PatternFill("solid", fgColor=BLUE)
        ws[ref].alignment = Alignment(horizontal="center", wrap_text=True)
    if not ws["A4"].value and sites:
        ws["A4"] = sites[0]
    ws["A4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    headers = {
        "D7": "Date/heure calculée",
        "E7": "Code horotarifaire",
        "F7": "Plage horaire",
        "G7": "Saison",
        "H7": "Tarif retenu",
        "I7": "Puissance souscrite [kVA]",
        "J7": "Dépassement puissance",
        "K7": "Ecart vs puissance souscrite [kW]",
        "L7": "Date référence contrat",
    }
    for ref, value in headers.items():
        ws[ref] = value
        ws[ref].font = Font(bold=True, color=WHITE)
        ws[ref].fill = PatternFill("solid", fgColor=GREEN)
        ws[ref].alignment = Alignment(horizontal="center", wrap_text=True)

    last_row = ws.max_row
    site_range = "Sites!$A$2:$A$5000"
    date_range = "Sites!$E$2:$E$5000"
    tarif_range = "Sites!$F$2:$F$5000"
    power_table = "Sites!$G$2:$R$5000"
    power_headers = "Sites!$G$1:$R$1"

    for row in range(8, last_row + 1):
        dt = f"$D{row}"
        tarif = f"$H{row}"
        code = f"$E{row}"
        ws.cell(row, 4).value = f'=IF($A{row}="","",DATE(VALUE(MID($A{row},7,4)),VALUE(MID($A{row},4,2)),VALUE(LEFT($A{row},2)))+TIME(VALUE(MID($A{row},12,2)),VALUE(MID($A{row},15,2)),0))'
        ws.cell(row, 12).value = f'=IFERROR(LOOKUP(2,1/(($A$4={site_range})*({date_range}<={dt})),{date_range}),"")'
        ws.cell(row, 8).value = f'=IFERROR(LOOKUP(2,1/(($A$4={site_range})*({date_range}<={dt})),{tarif_range}),"")'
        ws.cell(row, 5).value = (
            f'=IF({dt}="","",'
            f'IF({tarif}="Vert TE",'
            f'IF(OR(MONTH({dt})>=10,MONTH({dt})<=3),'
            f'IF(WEEKDAY({dt},2)>=6,IF(MOD({dt},1)<TIME(16,0,0),"HCSH","HPSH"),IF(MOD({dt},1)<TIME(8,0,0),"HCSH",IF(MOD({dt},1)<TIME(18,0,0),"HPSH",IF(MOD({dt},1)<TIME(22,0,0),"PO","HPSH")))),'
            f'IF(WEEKDAY({dt},2)>=6,"HCSB",IF(AND(MOD({dt},1)>=TIME(18,0,0),MOD({dt},1)<TIME(22,0,0)),"HPSB","HCSB"))),'
            f'IF({tarif}="Vert",'
            f'IF(OR(MOD({dt},1)>=TIME(22,30,0),MOD({dt},1)<TIME(6,30,0)),IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"HCH","HCE"),'
            f'IF(WEEKDAY({dt},2)>=6,IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"HPH","HPE"),'
            f'IF(OR(AND(MOD({dt},1)>=TIME(9,0,0),MOD({dt},1)<TIME(12,30,0)),AND(MOD({dt},1)>=TIME(19,0,0),MOD({dt},1)<TIME(20,30,0))),"PO",IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"HPH","HPE")))),'
            f'IF(OR(MOD({dt},1)>=TIME(22,0,0),MOD({dt},1)<TIME(6,0,0)),"HC","HP"))))'
        )
        ws.cell(row, 6).value = (
            f'=SWITCH({code},"PO","Pointe","HPSH","Saison Haute - Heures Pleines","HCSH","Saison Haute - Heures Creuses",'
            f'"HPSB","Saison Basse - Heures Pleines","HCSB","Saison Basse - Heures Creuses","HPE","Eté - Heures Pleines",'
            f'"HCE","Eté - Heures Creuses","HPH","Hiver - Heures Pleines","HCH","Hiver - Heures Creuses","HP","Heures Pleines","HC","Heures Creuses","BASE","Base","")'
        )
        ws.cell(row, 7).value = (
            f'=IF({dt}="","",IF({tarif}="Vert TE",IF(OR(MONTH({dt})>=10,MONTH({dt})<=3),"Saison Haute","Saison Basse"),'
            f'IF({tarif}="Vert",IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"Hiver austral","Eté austral"),"Toute l\'année")))'
        )
        row_lookup = f"LOOKUP(2,1/(($A$4={site_range})*({date_range}<={dt})),ROW({date_range})-ROW(Sites!$E$2)+1)"
        ws.cell(row, 9).value = f'=IFERROR(INDEX({power_table},{row_lookup},MATCH("P_"&{code},{power_headers},0)),"")'
        ws.cell(row, 10).value = f'=IF(OR($B{row}="",$I{row}=""),"",IF($B{row}>$I{row},"Oui","Non"))'
        ws.cell(row, 11).value = f'=IF(OR($B{row}="",$I{row}=""),"",$B{row}-$I{row})'
        for col in (4, 12):
            ws.cell(row, col).number_format = "dd/mm/yyyy hh:mm"
        for col in (9, 11):
            ws.cell(row, col).number_format = "0.0"

    ws["B4"] = '=IFERROR(INDEX(I:I,AGGREGATE(15,6,ROW($I$8:$I$60000)/($I$8:$I$60000<>""),1)),"")'
    ws["C4"] = '=IFERROR(INDEX(H:H,AGGREGATE(15,6,ROW($H$8:$H$60000)/($H$8:$H$60000<>""),1)),"")'
    for ref in ("B4", "C4"):
        ws[ref].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws[ref].alignment = Alignment(horizontal="center")

    for col, width in {
        "A": 24, "B": 20, "C": 18, "D": 20, "E": 18, "F": 32,
        "G": 18, "H": 18, "I": 22, "J": 20, "K": 26, "L": 20,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False
    thin = Side(style="hair", color="DDDDDD")
    for row in ws.iter_rows(min_row=1, max_row=min(last_row, 250), max_col=12):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def main() -> None:
    wb = load_workbook(WORKBOOK)
    sites_ws = wb["Sites"]
    sites = enhance_sites(sites_ws)
    treatments_ws = wb["Taitements"]
    reset_site_validation(treatments_ws, len(sites))
    add_treatments_formulas(treatments_ws, sites)
    wb.save(WORKBOOK)

    check = load_workbook(WORKBOOK, data_only=False, read_only=True)
    ws = check["Taitements"]
    assert ws["E7"].value == "Code horotarifaire"
    assert "LOOKUP" in str(ws["I8"].value)
    assert check["Sites"]["R1"].value == "P_HCSB"
    print(WORKBOOK)
    print({"sites_uniques": len(sites), "lignes_10min": ws.max_row - 7, "site_selectionne": ws["A4"].value})


if __name__ == "__main__":
    main()
