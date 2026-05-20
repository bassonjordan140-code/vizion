from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"

BLUE = "1F4E78"
GREEN = "70AD47"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"


def parse_power_text(value) -> dict[str, float | None]:
    if value is None:
        return {}
    text = str(value).strip()
    pairs = re.findall(r"([A-Z]+)\s*:\s*(\d+(?:[,.]\d+)?)", text, flags=re.I)
    if pairs:
        return {code.upper(): float(num.replace(",", ".")) for code, num in pairs}
    match = re.search(r"\d+(?:[,.]\d+)?", text)
    if not match:
        return {}
    value = float(match.group(0).replace(",", "."))
    return {code: value for code in ["BASE", "HP", "HC", "PO", "HPE", "HCE", "HPH", "HCH", "HPSH", "HCSH", "HPSB", "HCSB"]}


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            pass
    return value


def normalize_tarif(value) -> str:
    text = str(value or "").strip().lower()
    if "vert" in text and ("te" in text or "transition" in text):
        return "Vert TE"
    if "vert" in text:
        return "Vert"
    if "bleu plus" in text or "bleu+" in text:
        return "Bleu Plus"
    if "bleu" in text:
        return "Bleu"
    return str(value or "").strip()


def rebuild_sites_helpers(ws) -> list[str]:
    headers = [
        "site", "puissance_souscrite_kVA", "periode_fin_ref", "tarif_retenu",
        "P_BASE", "P_HP", "P_HC", "P_PO", "P_HPE", "P_HCE", "P_HPH", "P_HCH",
        "P_HPSH", "P_HCSH", "P_HPSB", "P_HCSB", None, "liste_sites",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
        if header:
            ws.cell(1, col).font = Font(bold=True, color=WHITE)
            ws.cell(1, col).fill = PatternFill("solid", fgColor=BLUE)
            ws.cell(1, col).alignment = Alignment(horizontal="center", wrap_text=True)

    seen = set()
    unique_sites = []
    for row in range(2, ws.max_row + 1):
        site = ws.cell(row, 1).value
        if site and site not in seen:
            unique_sites.append(str(site))
            seen.add(site)

        ws.cell(row, 3).value = parse_date(ws.cell(row, 3).value)
        ws.cell(row, 3).number_format = "yyyy-mm-dd"
        ws.cell(row, 4).value = normalize_tarif(ws.cell(row, 4).value)

        powers = parse_power_text(ws.cell(row, 2).value)
        for offset, code in enumerate(["BASE", "HP", "HC", "PO", "HPE", "HCE", "HPH", "HCH", "HPSH", "HCSH", "HPSB", "HCSB"], start=5):
            ws.cell(row, offset).value = powers.get(code) or powers.get("BASE")

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 18).value = None
    for row, site in enumerate(unique_sites, start=2):
        ws.cell(row, 18).value = site

    for col, width in {
        "A": 30, "B": 52, "C": 14, "D": 16, "E": 12, "F": 12, "G": 12, "H": 12,
        "I": 12, "J": 12, "K": 12, "L": 12, "M": 12, "N": 12, "O": 12, "P": 12,
        "R": 30,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    return unique_sites


def formula_code(row: int) -> str:
    dt = f"$D{row}"
    tarif = f"$H{row}"
    return (
        f'=IF({dt}="","",'
        f'IF({tarif}="Vert TE",'
        f'IF(OR(MONTH({dt})>=10,MONTH({dt})<=3),'
        f'IF(WEEKDAY({dt},2)>=6,IF(MOD({dt},1)<TIME(16,0,0),"HCSH","HPSH"),'
        f'IF(MOD({dt},1)<TIME(8,0,0),"HCSH",IF(MOD({dt},1)<TIME(18,0,0),"HPSH",IF(MOD({dt},1)<TIME(22,0,0),"PO","HPSH")))),'
        f'IF(WEEKDAY({dt},2)>=6,"HCSB",IF(AND(MOD({dt},1)>=TIME(18,0,0),MOD({dt},1)<TIME(22,0,0)),"HPSB","HCSB"))),'
        f'IF({tarif}="Vert",'
        f'IF(OR(MOD({dt},1)>=TIME(22,30,0),MOD({dt},1)<TIME(6,30,0)),IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"HCH","HCE"),'
        f'IF(WEEKDAY({dt},2)>=6,IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"HPH","HPE"),'
        f'IF(OR(AND(MOD({dt},1)>=TIME(9,0,0),MOD({dt},1)<TIME(12,30,0)),AND(MOD({dt},1)>=TIME(19,0,0),MOD({dt},1)<TIME(20,30,0))),"PO",IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"HPH","HPE")))),'
        f'IF(OR(MOD({dt},1)>=TIME(22,0,0),MOD({dt},1)<TIME(6,0,0)),"HC","HP"))))'
    )


def formula_label(row: int) -> str:
    code = f"$E{row}"
    return (
        f'=IF({code}="PO","Pointe",'
        f'IF({code}="HPSH","Saison Haute - Heures Pleines",'
        f'IF({code}="HCSH","Saison Haute - Heures Creuses",'
        f'IF({code}="HPSB","Saison Basse - Heures Pleines",'
        f'IF({code}="HCSB","Saison Basse - Heures Creuses",'
        f'IF({code}="HPE","Eté - Heures Pleines",'
        f'IF({code}="HCE","Eté - Heures Creuses",'
        f'IF({code}="HPH","Hiver - Heures Pleines",'
        f'IF({code}="HCH","Hiver - Heures Creuses",'
        f'IF({code}="HP","Heures Pleines",'
        f'IF({code}="HC","Heures Creuses",'
        f'IF({code}="BASE","Base",""))))))))))))'
    )


def repair_treatments(ws, unique_site_count: int) -> None:
    ws.data_validations.dataValidation = []
    validation = DataValidation(type="list", formula1=f"=Sites!$R$2:$R${unique_site_count + 1}", allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(ws["A4"])

    headers = {
        "A3": "Site sélectionné",
        "B3": "Puissance souscrite 1er pas [kVA]",
        "C3": "Tarif 1er pas",
        "D7": "Date/heure calculée",
        "E7": "Code horotarifaire",
        "F7": "Plage horaire",
        "G7": "Saison",
        "H7": "Tarif retenu",
        "I7": "Puissance souscrite [kVA]",
        "J7": "Dépassement puissance",
        "K7": "Ecart vs puissance souscrite [kW]",
        "L7": "Date référence contrat",
    }
    for ref, value in headers.items():
        ws[ref].value = value
        ws[ref].font = Font(bold=True, color=WHITE)
        ws[ref].fill = PatternFill("solid", fgColor=BLUE if ref[1:] == "3" else GREEN)
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["B4"] = '=IFERROR(INDEX(I:I,AGGREGATE(15,6,ROW($I$8:$I$60000)/($I$8:$I$60000<>""),1)),"")'
    ws["C4"] = '=IFERROR(INDEX(H:H,AGGREGATE(15,6,ROW($H$8:$H$60000)/($H$8:$H$60000<>""),1)),"")'
    for ref in ("B4", "C4"):
        ws[ref].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws[ref].alignment = Alignment(horizontal="center")

    site_range = "Sites!$A$2:$A$5000"
    date_range = "Sites!$C$2:$C$5000"
    tarif_range = "Sites!$D$2:$D$5000"
    power_table = "Sites!$E$2:$P$5000"
    power_headers = "Sites!$E$1:$P$1"

    for row in range(8, ws.max_row + 1):
        dt = f"$D{row}"
        code = f"$E{row}"
        row_lookup = f"LOOKUP(2,1/(($A$4={site_range})*({date_range}<={dt})),ROW({date_range})-ROW(Sites!$C$2)+1)"
        ws.cell(row, 4).value = f'=IF($A{row}="","",DATE(VALUE(MID($A{row},7,4)),VALUE(MID($A{row},4,2)),VALUE(LEFT($A{row},2)))+TIME(VALUE(MID($A{row},12,2)),VALUE(MID($A{row},15,2)),0))'
        ws.cell(row, 5).value = formula_code(row)
        ws.cell(row, 6).value = formula_label(row)
        ws.cell(row, 7).value = f'=IF({dt}="","",IF($H{row}="Vert TE",IF(OR(MONTH({dt})>=10,MONTH({dt})<=3),"Saison Haute","Saison Basse"),IF($H{row}="Vert",IF(AND(MONTH({dt})>=5,MONTH({dt})<=9),"Hiver austral","Eté austral"),"Toute l\'année")))'
        ws.cell(row, 8).value = f'=IFERROR(LOOKUP(2,1/(($A$4={site_range})*({date_range}<={dt})),{tarif_range}),"")'
        ws.cell(row, 9).value = f'=IFERROR(INDEX({power_table},{row_lookup},MATCH("P_"&{code},{power_headers},0)),"")'
        ws.cell(row, 10).value = f'=IF(OR($B{row}="",$I{row}=""),"",IF($B{row}>$I{row},"Oui","Non"))'
        ws.cell(row, 11).value = f'=IF(OR($B{row}="",$I{row}=""),"",$B{row}-$I{row})'
        ws.cell(row, 12).value = f'=IFERROR(LOOKUP(2,1/(($A$4={site_range})*({date_range}<={dt})),{date_range}),"")'
        ws.cell(row, 4).number_format = "dd/mm/yyyy hh:mm"
        ws.cell(row, 12).number_format = "yyyy-mm-dd"

    for col, width in {
        "A": 24, "B": 24, "C": 16, "D": 20, "E": 18, "F": 34,
        "G": 18, "H": 16, "I": 22, "J": 20, "K": 28, "L": 18,
    }.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False
    thin = Side(style="hair", color="DDDDDD")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 250), max_col=12):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def main() -> None:
    wb = load_workbook(WORKBOOK, keep_links=False)
    sites = wb["Sites"]
    unique_sites = rebuild_sites_helpers(sites)
    repair_treatments(wb["Taitements"], len(unique_sites))
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(WORKBOOK)

    check = load_workbook(WORKBOOK, data_only=False, read_only=True, keep_links=False)
    ws = check["Taitements"]
    assert "#REF!" not in str(ws["I8"].value)
    assert "_xludf" not in str(ws["F8"].value)
    assert check["Sites"]["R1"].value == "liste_sites"
    print(WORKBOOK)
    print({"sites_uniques": len(unique_sites), "lignes_10min": ws.max_row - 7, "site_selectionne": ws["A4"].value})


if __name__ == "__main__":
    main()
