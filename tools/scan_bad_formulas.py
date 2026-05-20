from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True, keep_links=False)
    bad = []
    for name in ["Taitements", "Sites"]:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                if "#REF!" in value or "_xludf" in value:
                    bad.append((name, cell.coordinate, value[:160]))
    print("bad_count", len(bad))
    for item in bad[:20]:
        print(item)


if __name__ == "__main__":
    main()
