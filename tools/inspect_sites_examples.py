from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "grille_tarifaire" / "grille_tarifaire_edf_reunion_2020_2026.xlsx"


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb["Sites"]
    for needle in ["Omega", "Aurar Omega", "Aurar St Benoit", "Aurar Siege"]:
        print(f"--- {needle}")
        count = 0
        for row in range(1, ws.max_row + 1):
            site = ws.cell(row, 1).value
            if site and needle.lower() in str(site).lower():
                print(row, [ws.cell(row, col).value for col in range(1, min(ws.max_column, 8) + 1)])
                count += 1
                if count >= 20:
                    break


if __name__ == "__main__":
    main()
