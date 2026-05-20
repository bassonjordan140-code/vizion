from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "outputs" / "grille_tarifaire_work" / "tariff_data.json"
OUT_DIR = ROOT / "outputs" / "grille_tarifaire"
OUT_FILE = OUT_DIR / "grille_tarifaire_edf_reunion_2020_2026.xlsx"

BLUE = "1F4E78"
GREEN = "70AD47"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
GREY = "F2F2F2"
WHITE = "FFFFFF"


def dedupe_rows(rows: list[dict]) -> list[dict]:
    grouped = {}
    sources = defaultdict(list)
    for row in rows:
        key = tuple((k, row.get(k)) for k in sorted(row.keys()) if k != "source_pdf")
        if key not in grouped:
            grouped[key] = {k: v for k, v in row.items() if k != "source_pdf"}
        if row.get("source_pdf") not in sources[key]:
            sources[key].append(row.get("source_pdf"))
    result = []
    for key, row in grouped.items():
        row["sources_pdf"] = " | ".join(s for s in sources[key] if s)
        row["nb_sources_identiques"] = len([s for s in sources[key] if s])
        result.append(row)
    return sorted(result, key=lambda r: (r.get("date_application") or "", r.get("tarif") or "", r.get("option") or "", str(r.get("puissance_souscrite") or ""), r.get("periode_prix") or ""))


def write_sheet(ws, headers: list[str], rows: list[dict], title: str | None = None) -> None:
    start_row = 1
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, color=WHITE, size=14)
        ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        start_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header))
    if rows:
        ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
        table = Table(displayName=f"Table_{ws.title.replace(' ', '_').replace('-', '_')}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{max(start_row + len(rows), start_row)}"
    for col_idx, header in enumerate(headers, start=1):
        values = [str(header)] + ["" if row.get(header) is None else str(row.get(header)) for row in rows[:200]]
        width = min(max(max(len(v) for v in values) + 2, 10), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="DDDDDD"))


def build() -> None:
    payload = json.loads(DATA.read_text(encoding="utf-8"))
    price_rows = dedupe_rows(payload["prix"])
    time_rows = dedupe_rows(payload["plages_horaires"])
    sources = sorted(payload["sources"], key=lambda r: (r.get("date_application") or "", r.get("tarif") or "", r.get("source_pdf") or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese"
    ws["A1"] = "Grille tarifaire EDF Réunion - données consolidées"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A1:F1")
    summary = [
        ("Période couverte par les fichiers", "2020-08-01 à 2026-02-01"),
        ("Sources PDF recensées", len(sources)),
        ("Lignes de prix consolidées", len(price_rows)),
        ("Lignes de plages horaires consolidées", len(time_rows)),
        ("PDF sans texte exploitable", sum(1 for s in sources if s["caracteres_extraits"] == 0)),
        ("Note", "Les PDF sans calque texte sont listés dans Sources avec statut OCR requis."),
    ]
    for idx, (label, value) in enumerate(summary, start=3):
        ws.cell(idx, 1, label).font = Font(bold=True)
        ws.cell(idx, 2, value)
    ws["A11"] = "Onglets"
    ws["A11"].font = Font(bold=True, color=WHITE)
    ws["A11"].fill = PatternFill("solid", fgColor=GREEN)
    tabs = [
        ("Prix", "Prix d'énergie et abonnements extraits, dédoublonnés par date/tarif/option/période."),
        ("Plages horaires", "Créneaux heures pleines, heures creuses, pointes et saisons."),
        ("Sources", "Inventaire complet des PDF, statut d'extraction et nombre de lignes reconnues."),
    ]
    for idx, (tab, desc) in enumerate(tabs, start=12):
        ws.cell(idx, 1, tab).font = Font(bold=True)
        ws.cell(idx, 2, desc)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="DDDDDD"))

    prix_headers = [
        "date_application", "annee", "tarif", "option", "segment_version", "puissance_souscrite",
        "reglage_disjoncteur_A", "abonnement", "unite_abonnement", "majoration_abonnement_EUR_kVA_an",
        "periode_prix", "prix_energie_cEUR_kWh", "coefficient_puissance_reduite",
        "depassement_EUR_kW", "energie_reactive_cEUR_kVArh", "nb_sources_identiques", "sources_pdf",
    ]
    write_sheet(wb.create_sheet("Prix"), prix_headers, price_rows, "Prix consolidés")

    plages_headers = ["date_application", "tarif", "option", "saison", "periode", "plage_horaire", "nb_sources_identiques", "sources_pdf"]
    write_sheet(wb.create_sheet("Plages horaires"), plages_headers, time_rows, "Plages horaires")

    source_headers = ["date_application", "tarif", "source_pdf", "pages", "caracteres_extraits", "lignes_prix_extraites", "statut"]
    write_sheet(wb.create_sheet("Sources"), source_headers, sources, "Sources PDF")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)

    check = load_workbook(OUT_FILE, data_only=False)
    assert set(check.sheetnames) == {"Synthese", "Prix", "Plages horaires", "Sources"}
    assert check["Prix"].max_row > 100
    assert check["Sources"].max_row == len(sources) + 3
    print(OUT_FILE)
    print({"prix_consolides": len(price_rows), "plages_horaires": len(time_rows), "sources": len(sources)})


if __name__ == "__main__":
    build()
